# README:
# Step 1: Copy source workbook file path into load_workbook argument, line 20;
# please make sure source workbook and python file are in same folder
# Step 2: Make sure source workbook cells are values only and not formulas;
# you can ensure this by copying all the values over to a new sheet and pasting Values
# Step 3: Set file name for destination file within the string in dest_filename variable, line 25
# Step 4: Run program and input all variables as requested ("recommended" variables are for Input_Dataset only)
# Note: If destination file is already open, please close it before running the program,
# otherwise program will not be able to access destination file

# Imports relevant modules
from openpyxl import Workbook, load_workbook
from numpy import array
from statistics import mean
from scipy.signal import argrelmin
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle

# Loads source workbook [note: please make sure workbook is values only; no formulas]
wb1 = load_workbook(filename = 'Input_Dataset.xlsx')
sheet = wb1.active

# Loads destination workbook
wb2 = Workbook()
dest_filename = 'Output_File.xlsx'

# Loads a new sheet for the destination workbook
ws1 = wb2.active
ws1.title = 'Tm data'

# Sets number of compounds, number of titrations, and number of rows
compound_count = int(input("Enter number of compounds to be tested (For Input_Dataset, enter 3): "))
titration_count = int(input("Enter number of titrations used (For Input_dataset, enter 5): "))
row_count = int(input("Enter number of data points used for first derivative plot (For Input_Dataset, enter 121): "))
max_Tm_count = 5

conc_list = []
for conc_ind in range(titration_count):
    conc = round(float(input('Enter concentration {} of {}, in uM (lowest to highest, excluding apo): '.format(conc_ind + 1, titration_count))),2)
    conc_list.append(conc)
    # Allows user to fill in concentrations until required number of titrations is reached
conc_list.insert(0, round(float(0), 2))
conc_count = len(conc_list)

tem_min = float(input('Enter temperature lower bound for Tm determination, in °C (default 55): '))
tem_max = float(input('Enter temperature upper bound for Tm determination, in °C (default 80): '))
resp_min = float(input('Enter response lower bound for Tm determination, in response units (default -5.0): '))
resp_max = float(input('Enter response upper bound for Tm determination, in response units (default -0.5): '))

# Sets boundary conditions for valley picking
(tem_width, resp_height) = (tem_max - tem_min, resp_max - resp_min)

#Extracts temperature values from first column in sheet
tem_value_list = []
for tem_value in sheet.iter_cols(min_row = 2, max_row = row_count + 1, max_col = 1, values_only = True):
    tem_value_list = list(tem_value)
tem_value_array = array(tem_value_list)

# Sorts apo and non-apo column indices
apo_avg_resp_list = []
apo_col_indices = []
non_apo_col_indices = []
column_count = int(compound_count * (titration_count + 1))   # 3 comps * (5 titrs per comp + 1 apo per comp) = 18

for col_ind in range(column_count):     # 16 loops for 16 columns
    data_col_ind = (col_ind + 1) * 2
    column_name = str(sheet.cell(row = 1, column = data_col_ind).value)
    if "APO" in column_name or "apo" in column_name:
        apo_col_indices.append(data_col_ind)
    else:
        non_apo_col_indices.append(data_col_ind)

# Compresses all apo columns into one single apo column
# containing mean of all apo response values at each temp
for row_ind in range(2, row_count + 2):    # 128 loops for 128 temperature values
    apo_values_by_tem = []
    # when adding cells using 'for' loops, always put empty list before loop
    for col_ind in apo_col_indices:        # 16 loops for 16 columns
        apo_resp = sheet.cell(row = row_ind, column = col_ind).value
        apo_values_by_tem.append(apo_resp)
    apo_mean_by_tem = float(mean(apo_values_by_tem))
    apo_avg_resp_list.append(apo_mean_by_tem)

# Makes a list of compound names
compound_names = []
for col_ind in non_apo_col_indices:                 # 320 loops for 320 non-apo columns
    if col_ind % (titration_count * 2) == 2:        # Takes the first of every 5 columns
        column_name = sheet.cell(row = 1, column = col_ind).value
        compound_names.append(column_name[:12])     # First 12 characters of string are the compound name

# Produces a collection of response value lists for each compound in each titration
resp_values_total = []

# Selects the column index to pick from non_apo_col_indices, if concentration is 0 then returns "APO"
column_select = lambda comp_ind, conc_ind: (comp_ind * titration_count + conc_ind - 1) if conc_ind != 0 else "APO"

for comp_ind in range(compound_count):          # 3 loops for 3 compounds
    for conc_ind in range(conc_count):          # 6 loops for 6 concentrations
       
        compound = compound_names[comp_ind]                     # Compound Name
        conc = conc_list[conc_ind]                              # Concentration in uM
        chosen_col_ind = column_select(comp_ind, conc_ind)      # See note for column_select above
        chosen_col = non_apo_col_indices[chosen_col_ind] if chosen_col_ind != "APO" else "APO"
        resp_column = []
       
        if chosen_col != "APO":
            for resp in sheet.iter_cols(min_row = 2, max_row = row_count + 1, min_col = chosen_col, max_col = chosen_col, values_only = True):
                resp_column = list(resp)
        else:
            resp_column = apo_avg_resp_list
       
        resp_values_total.append(resp_column)

# Produces a collection of Tm's for each column in resp_values_total, and deposits them in Excel file
bounded_indices_list = []

for comp_ind in range(compound_count):          # 3 loops for 3 compounds
    for conc_ind in range(conc_count):          # 6 loops for 6 concentrations

        compound = compound_names[comp_ind]             # Compound Name
        conc = conc_list[conc_ind]                      # Concentration in uM
        total_ind = comp_ind * conc_count + conc_ind    # Index of column in resp_values_total
        datapt_list = resp_values_total[total_ind]
        datapt_array = array(datapt_list)
       
        # Takes all local minimum y-values over the entire curve and outputs an array
        # order determines minimum compared to how many points on either side
        minima_indices = argrelmin(datapt_array, order = 5)[0]
        minima_tem = tem_value_array[minima_indices]        # Tm of valley
        minima_resp = datapt_array[minima_indices]          # Y-value of valley

        # Filters out all minima and Tm's which satisfy the temperature and response value boundaries
        bounds_tem = lambda i: minima_tem[i] > tem_min and minima_tem[i] < tem_max
        bounds_resp = lambda i: minima_resp[i] > resp_min and minima_resp[i] < resp_max
        bounds_both = lambda i: bounds_tem(i) and bounds_resp(i)
        # Don't forget the 'i' argument in all of the boolean functions above, it is essential
        bounded_indices = [minima_indices[i] for i in range(len(minima_indices)) if bounds_both(i) == True]
        bounded_indices_list.append(bounded_indices)
        Tm_values_by_col = [tem_value_list[i] for i in bounded_indices]

        # Enters compound names, concentrations and Tm values in destination table
        row_ind = total_ind + 2
        for i in range(len(Tm_values_by_col)):
            for col_ind, dest_value in [[1, compound], [2, conc], [i + 3, Tm_values_by_col[i]]]:
                ws1.cell(row = row_ind, column = col_ind, value = dest_value)

# Inputs column names in destination sheet
column_names = ['Compound', 'Concentration (uM)']
for i in range(max_Tm_count):
    column_names.append('Tm{}'.format(i + 1))   # Assigns names Tm1, Tm2, etc.
for col in range(len(column_names)):
    ws1.cell(row = 1, column = col + 1, value = column_names[col])

# Saves the destination workbook
wb2.save(filename = dest_filename)

# Allows user to open destination file and begin plotting at their own pace
print("Destination file ready.")
next_step = input("Please open destination file, and then hit enter to begin plotting.")

for comp_ind in range(compound_count):          # 64 loops for 64 compounds
    for conc_ind in range(conc_count):          # 6 loops for 6 concentrations

        compound = compound_names[comp_ind]             # Compound Name
        conc = conc_list[conc_ind]                      # Concentration in uM
        total_ind = comp_ind * conc_count + conc_ind    # Index of column in resp_values_total
        bounded_indices = bounded_indices_list[total_ind]
        datapt_list = resp_values_total[total_ind]
        datapt_array = array(datapt_list)

        plt.plot(tem_value_array, datapt_array, label = '{} uM'.format(conc))           # Curve
        plt.plot(tem_value_array[bounded_indices], datapt_array[bounded_indices], "x")  # Tm valleys
        plt.xlabel("Temperature")                                                       # x-axis label
        plt.ylabel("-df/dt")                                                            # y-axis label
        plt.title("({} of {}) {}".format(comp_ind + 1, compound_count, compound))       # Graph title
        plt.gca().add_patch(Rectangle((tem_min, resp_min), tem_width, resp_height, edgecolor = 'black', facecolor = 'none', linewidth = 2, linestyle = 'dotted'))

    plt.legend()    # Shows a common legend for all curves in each compound
    plt.show()      # Shows all plots for each compound

print("All plots generated. End of program.")
end = input("Hit enter to terminate.")